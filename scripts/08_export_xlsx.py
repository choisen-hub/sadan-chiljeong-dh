"""
08_export_xlsx.py

annotated jsonl을 분석용 xlsx로 export.

시트 구성:
  - summary           : 전체 통계 요약
  - all_sentences     : 전체 sentence (29K rows × 14 cols, 분석 시 필터 기준)
  - li_sentences      : 理 포함 필터 (주 분석 편의용 — 대조군 분석은
                         all_sentences에서 동적으로 추출)

스타일:
  - 헤더: 굵게 + 진한 파란 배경 + 흰 글씨
  - frozen panes (헤더 고정)
  - 컬럼 너비: 본문 컬럼은 넓게, 식별자는 좁게

입력:  data/intermediate/zhuzi_sentences_annotated.jsonl
출력:  data/final/zhuzi_sentences.xlsx

사용법:
  python3 scripts/08_export_xlsx.py
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

PROJECT_ROOT = Path(__file__).resolve().parent.parent
INPUT = PROJECT_ROOT / "data" / "intermediate" / "zhuzi_sentences_annotated.jsonl"
OUT = PROJECT_ROOT / "data" / "final" / "zhuzi_sentences.xlsx"

# 분석 시 자주 쓰는 순서로 컬럼 정렬
COLUMNS = [
    "sentence_id",
    "juan_num", "juan_label",
    "paragraph_idx", "sentence_idx", "utterance_id",
    "text_punctuated", "text_plain",
    "char_count",
    "has_li", "has_qi", "li_qi_category",
    "byline", "headings",
]

# 컬럼별 너비 (Excel 단위)
COL_WIDTHS = {
    "sentence_id":     12,
    "juan_num":         8,
    "juan_label":      12,
    "paragraph_idx":   10,
    "sentence_idx":    10,
    "utterance_id":    14,
    "text_punctuated": 60,
    "text_plain":      50,
    "char_count":       9,
    "has_li":           7,
    "has_qi":           7,
    "li_qi_category":  13,
    "byline":          14,
    "headings":        35,
    "항목":            30,
    "값":              12,
}


def load_records(path: Path) -> list[dict]:
    """jsonl 로드 후 dict/list 컬럼은 셀에 넣기 위해 string화."""
    out: list[dict] = []
    with path.open(encoding="utf-8") as f:
        for line in f:
            r = json.loads(line)
            r["byline"] = (
                json.dumps(r["byline"], ensure_ascii=False) if r.get("byline") else ""
            )
            r["headings"] = json.dumps(r.get("headings", []), ensure_ascii=False)
            out.append(r)
    return out


def style_sheet(ws) -> None:
    """헤더 스타일 + 컬럼 너비 + frozen panes."""
    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    header_fill = PatternFill("solid", start_color="305496")
    header_align = Alignment(horizontal="center", vertical="center")

    # 헤더 row 스타일
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # 컬럼 너비
    for col_idx, cell in enumerate(ws[1], start=1):
        col_name = cell.value
        width = COL_WIDTHS.get(col_name, 14)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    # 헤더 고정
    ws.freeze_panes = "A2"


def main() -> None:
    if not INPUT.exists():
        print(f"[ERROR] 입력 없음: {INPUT}", file=sys.stderr)
        print("        먼저 07_annotate.py 실행 필요", file=sys.stderr)
        sys.exit(1)

    print(f"[LOAD] {INPUT.name}")
    records = load_records(INPUT)
    df = pd.DataFrame(records)[COLUMNS]
    print(f"       {len(df):,} rows × {len(df.columns)} cols")

    # 시트별 슬라이싱 (분석 대상은 전수, 편의용 li_sentences만)
    df_li = df[df["has_li"]].copy()

    # summary 시트 데이터
    n_total = len(df)
    n_li = int(df["has_li"].sum())
    n_qi = int(df["has_qi"].sum())
    n_both = int((df["has_li"] & df["has_qi"]).sum())
    n_li_only = int((df["has_li"] & ~df["has_qi"]).sum())
    n_qi_only = int((~df["has_li"] & df["has_qi"]).sum())
    n_neither = int((~df["has_li"] & ~df["has_qi"]).sum())

    summary_rows = [
        ("[전체] 총 sentence 수",      n_total),
        ("[전체] 理 포함",             n_li),
        ("[전체] 氣 포함",             n_qi),
        ("[전체] 理+氣 둘 다",         n_both),
        ("[전체] 理만 (氣 없음)",      n_li_only),
        ("[전체] 氣만 (理 없음)",      n_qi_only),
        ("[전체] 둘 다 없음",          n_neither),
        ("",                            ""),
        ("[메타] 평균 char_count",     round(df["char_count"].mean(), 1)),
        ("[메타] 중위 char_count",     int(df["char_count"].median())),
        ("[메타] 최댓값 char_count",   int(df["char_count"].max())),
        ("[메타] 최솟값 char_count",   int(df["char_count"].min())),
    ]
    df_summary = pd.DataFrame(summary_rows, columns=["항목", "값"])

    # ----- xlsx 작성 -----
    OUT.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="summary", index=False)
        df.to_excel(writer, sheet_name="all_sentences", index=False)
        df_li.to_excel(writer, sheet_name="li_sentences", index=False)

    # ----- formatting -----
    wb = load_workbook(OUT)
    for sheet_name in wb.sheetnames:
        style_sheet(wb[sheet_name])
    wb.save(OUT)

    print(f"\n[DONE] 저장: {OUT}")
    print(f"       시트:")
    print(f"         summary")
    print(f"         all_sentences      {n_total:>7,} rows")
    print(f"         li_sentences       {n_li:>7,} rows  (理 포함)")


if __name__ == "__main__":
    main()
