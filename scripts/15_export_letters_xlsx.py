#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
15_export_letters_xlsx.py — Export letters + sentences to xlsx for review.

Step 15 of Phase 0 pipeline. Builds a single Excel file with 3 sheets:

  Summary    corpus 통계 + 理/氣 분포 (퇴계 vs 율곡 대비)
  Letters    31편 letter list (raw_text + punctuated_text 모두 포함)
  Sentences  분절 + annotate 결과 문장들 (has_li/has_qi 컬럼 포함)

이전 14_export_letters_xlsx.py (v2.0) 와의 차이:
  - 입력: sentences.jsonl → sentences_annotated.jsonl (14_annotate_letters.py 산출)
  - Sentences sheet 에 has_li / has_qi / li_qi_category 컬럼 추가
  - Summary sheet 에 理/氣 카테고리 분포 (퇴계/율곡 분리) 추가

USAGE
=====

  python3 scripts/15_export_letters_xlsx.py


OUTPUT
======

  data/final/corpus_review.xlsx


DEPENDENCIES
============

  Python ≥ 3.9 + openpyxl
    pip install openpyxl
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
from collections import defaultdict
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Missing dependency: openpyxl\nInstall: pip install openpyxl")


DEFAULT_REPO_ROOT = Path(__file__).resolve().parent.parent

# Visual style
HEADER_FILL = PatternFill("solid", start_color="E0E0E0")
HEADER_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=13)
DEFAULT_FONT_NAME = "Arial"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _load_jsonl(path: Path) -> list[dict]:
    if not path.exists():
        sys.exit(f"Input not found: {path}\nRun upstream scripts first.")
    out = []
    with path.open(encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                out.append(json.loads(line))
    return out


def _style_header_row(ws, row_idx: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def _set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _apply_default_font(ws, font_name: str = DEFAULT_FONT_NAME) -> None:
    """Apply consistent font to all cells with content."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                existing = cell.font
                cell.font = Font(
                    name=font_name,
                    bold=existing.bold,
                    italic=existing.italic,
                    size=existing.size,
                    color=existing.color,
                )


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_summary_sheet(ws, letters: list[dict], sentences: list[dict]) -> None:
    ws.title = "Summary"

    # Aggregations
    n_letters = len(letters)
    n_sentences = len(sentences)
    total_chars_raw = sum(L["char_count_raw"] for L in letters)
    total_chars_plain = sum(L["char_count_plain"] for L in letters)
    total_chars_punc = sum(len(L.get("punctuated_text", "")) for L in letters)

    by_target: dict[str, list[dict]] = defaultdict(list)
    for L in letters:
        by_target[L["target_name"]].append(L)

    by_target_kwon: dict[tuple, list[dict]] = defaultdict(list)
    for L in letters:
        by_target_kwon[(L["target_name"], L["kwon"])].append(L)

    sents_by_target: dict[str, int] = defaultdict(int)
    for s in sentences:
        sents_by_target[s["target_name"]] += 1

    # target 별 理/氣 카테고리 분포 집계
    cat_by_target: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for s in sentences:
        cat_by_target[s["target_name"]][s.get("li_qi_category", "neither")] += 1

    # Title
    ws["A1"] = "사단칠정 corpus — Phase 0 추출 결과 (hanja.dev 표점 + 理/氣 annotate)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")

    # Overall stats
    row = 3
    ws.cell(row=row, column=1, value="총계").font = HEADER_FONT
    row += 1
    avg_sent_len = round(total_chars_plain / n_sentences, 2) if n_sentences else 0
    overall = [
        ("Letter 편수", n_letters),
        ("문장 수 (sentence)", n_sentences),
        ("총 글자 수 (한국문집총간 표점, raw)", total_chars_raw),
        ("총 글자 수 (백문, plain)", total_chars_plain),
        ("총 글자 수 (hanja.dev 표점)", total_chars_punc),
        ("평균 문장 길이 (백문)", avg_sent_len),
    ]
    for label, value in overall:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1

    # Target breakdown
    row += 2
    ws.cell(row=row, column=1, value="Target별 분포").font = HEADER_FONT
    row += 1
    headers = ["Target", "권", "편수", "자수 (raw)", "자수 (백문)", "문장 수"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=row, column=c, value=h)
    _style_header_row(ws, row, len(headers))
    row += 1

    for target in sorted(by_target.keys()):
        target_letters = by_target[target]
        kwons = sorted({L["kwon"] for L in target_letters})
        for k in kwons:
            kwon_letters = by_target_kwon[(target, k)]
            ws.cell(row=row, column=1, value=target)
            ws.cell(row=row, column=2, value=f"권{k}")
            ws.cell(row=row, column=3, value=len(kwon_letters))
            ws.cell(row=row, column=4, value=sum(L["char_count_raw"] for L in kwon_letters))
            ws.cell(row=row, column=5, value=sum(L["char_count_plain"] for L in kwon_letters))
            ws.cell(row=row, column=6, value="")
            row += 1
        ws.cell(row=row, column=1, value=target)
        ws.cell(row=row, column=2, value="합계").font = HEADER_FONT
        ws.cell(row=row, column=3, value=len(target_letters)).font = HEADER_FONT
        ws.cell(row=row, column=4,
                value=sum(L["char_count_raw"] for L in target_letters)).font = HEADER_FONT
        ws.cell(row=row, column=5,
                value=sum(L["char_count_plain"] for L in target_letters)).font = HEADER_FONT
        ws.cell(row=row, column=6, value=sents_by_target[target]).font = HEADER_FONT
        row += 2

    # 理/氣 카테고리 분포 (퇴계 vs 율곡 대비)
    row += 1
    ws.cell(row=row, column=1, value="理/氣 카테고리 분포 (target별)").font = HEADER_FONT
    row += 1
    cat_headers = ["Target", "category", "문장 수", "비율 (%)"]
    for c, h in enumerate(cat_headers, start=1):
        ws.cell(row=row, column=c, value=h)
    _style_header_row(ws, row, len(cat_headers))
    row += 1

    for target in sorted(cat_by_target.keys()):
        cnt = cat_by_target[target]
        sub_total = sum(cnt.values())
        for k in ["both", "li_only", "qi_only", "neither"]:
            n = cnt.get(k, 0)
            ws.cell(row=row, column=1, value=target)
            ws.cell(row=row, column=2, value=k)
            ws.cell(row=row, column=3, value=n)
            ws.cell(row=row, column=4,
                    value=round(n / max(sub_total, 1) * 100, 1))
            row += 1
        row += 1

    # Notes
    row += 1
    ws.cell(row=row, column=1, value="참고").font = HEADER_FONT
    row += 1
    notes = [
        "• 출처: 공공데이터포털 한국문집총간 release 2024-08-30 (data.go.kr/data/3074298)",
        "• 퇴계 측 22편 = 권16+권17 명언 letters (황준연(2009) 표 1의 12편 corpus의 superset)",
        "• 율곡 측 9편 = 권11 답성호원 letters (1572 壬申, 사단칠정 논변 본체)",
        "• 표점 부여: SikuRoBERTa-PUNC-AJD-KLC (한국고전번역원 AJD/KLC 표점 시스템)",
        "• 분절 기준: hanja.dev 출력의 종결자 (。？！?!) — 한 letter = 1 paragraph",
        "• 理/氣 annotate: sent_text_plain (백문) 기준, 표점에 우연히 동일 글자가 있어도 영향 없음",
        "• 자세한 판본 정보: docs/판본정보.md",
        "• 모델 호출 메타데이터: docs/letter_punctuation_provenance.md",
    ]
    for note in notes:
        ws.cell(row=row, column=1, value=note)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1

    _set_col_widths(ws, [38, 12, 10, 14, 14, 10])


def build_letters_sheet(ws, letters: list[dict]) -> None:
    ws.title = "Letters"

    headers = [
        "data_id", "munjip", "권", "seq", "제목", "원주(干支)",
        "추정연도", "발신측", "target",
        "raw 자수", "백문 자수", "표점 자수",
        "백문 첫 50자", "표점 첫 80자",
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    for L in letters:
        punc = L.get("punctuated_text", "")
        ws.append([
            L["data_id"],
            L["munjip"],
            L["kwon"],
            L["seq"],
            L["title"],
            L["annotation"] or "",
            L["inferred_year"] or "",
            L["sender_label"],
            L["target_name"],
            L["char_count_raw"],
            L["char_count_plain"],
            len(punc),
            L["body_first_50"],
            punc[:80],
        ])

    _set_col_widths(ws, [32, 8, 5, 5, 30, 22, 9, 8, 26, 10, 10, 10, 50, 80])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"


def build_sentences_sheet(ws, sentences: list[dict]) -> None:
    ws.title = "Sentences"

    headers = [
        "sentence_id", "source_id", "letter_id", "권", "letter_year", "발신측",
        "para", "sent#para", "sent#letter",
        "백문 자수", "has_li", "has_qi", "li_qi_category",
        "원문 (표점)", "백문",
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    for s in sentences:
        ws.append([
            s["sentence_id"],
            s.get("source_id", ""),
            s["letter_data_id"],
            s["kwon"],
            s["letter_year"] or "",
            s["sender_label"],
            s["para_idx"],
            s["sent_idx_in_para"],
            s["sent_idx_in_letter"],
            s["char_count_plain"],
            s.get("has_li", False),
            s.get("has_qi", False),
            s.get("li_qi_category", ""),
            s["sent_text"],
            s["sent_text_plain"],
        ])

    _set_col_widths(ws, [10, 36, 32, 5, 10, 8, 6, 8, 9, 10, 8, 8, 13, 60, 60])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--repo-root", type=Path, default=DEFAULT_REPO_ROOT)
    ap.add_argument("--letters", type=Path, default=None,
                    help="Override letters_punctuated.jsonl path")
    ap.add_argument("--sentences", type=Path, default=None,
                    help="Override sentences.jsonl path")
    ap.add_argument("--output", type=Path, default=None,
                    help="Override output xlsx path "
                         "(default: <repo>/data/final/corpus_review.xlsx)")
    ap.add_argument("--debug", action="store_true")
    args = ap.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.debug else logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
    )

    repo_root = args.repo_root.resolve()
    letters_path = args.letters or (repo_root / "data/processed/letters_punctuated.jsonl")
    sentences_path = args.sentences or (repo_root / "data/processed/sentences_annotated.jsonl")
    out_path = args.output or (repo_root / "data/final/corpus_review.xlsx")

    letters = _load_jsonl(letters_path)
    sentences = _load_jsonl(sentences_path)
    logging.info("Loaded %d letters, %d sentences", len(letters), len(sentences))

    wb = Workbook()
    build_summary_sheet(wb.active, letters, sentences)
    build_letters_sheet(wb.create_sheet("Letters"), letters)
    build_sentences_sheet(wb.create_sheet("Sentences"), sentences)

    for ws in wb.worksheets:
        _apply_default_font(ws)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    logging.info("Wrote xlsx → %s", out_path)
    logging.info("  3 sheets: Summary, Letters (%d rows), Sentences (%d rows)",
                 len(letters), len(sentences))


if __name__ == "__main__":
    main()
